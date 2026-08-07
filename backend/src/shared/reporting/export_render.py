"""Standard Reporting Framework — one tabular renderer used by every report
endpoint across every module. Owner directive: reports were each hand-
rolling their own filters/print/export, which is exactly how the Trial
Balance abnormal-balance bug slipped through undetected on other reports.
A single `ReportTable` -> PDF/Excel path means every future report inherits
correct print quality and export for free, instead of reinventing it.

PDF uses WeasyPrint (HTML/CSS -> PDF via Pango/Cairo), not a from-scratch
canvas library (e.g. reportlab): Pango does correct Arabic bidi/glyph-
joining out of the box, which a hand-rolled renderer would have to
reimplement and would likely get wrong for a bilingual Arabic-first ERP —
confirmed by a live test render (see docs/project-progress.md).
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from html import escape as h

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from weasyprint import HTML


@dataclass(frozen=True)
class ReportColumn:
    label: str
    align: str = "start"  # "start" | "end" — logical, flips with rtl


@dataclass(frozen=True)
class ReportTable:
    title: str
    company_name: str
    columns: list[ReportColumn]
    rows: list[list[str]]
    subtitle: str | None = None
    totals: list[str] | None = None
    highlight_rows: set[int] = field(default_factory=set)
    rtl: bool = True


def render_pdf(table: ReportTable) -> bytes:
    dir_attr = "rtl" if table.rtl else "ltr"

    def cell_align(col: ReportColumn) -> str:
        return "end" if col.align == "end" else "start"

    header_cells = "".join(
        f'<th style="text-align:{cell_align(c)}">{h(c.label)}</th>' for c in table.columns
    )

    def row_html(cells: list[str], *, bold: bool = False, highlight: bool = False) -> str:
        style = "font-weight:700;border-top:2px solid #333;" if bold else ""
        if highlight:
            style += "background:#fff8d6;"
        tds = "".join(
            f'<td style="text-align:{cell_align(table.columns[i])}">{h(str(v))}</td>'
            for i, v in enumerate(cells)
        )
        return f'<tr style="{style}">{tds}</tr>'

    body_rows = "".join(
        row_html(row, highlight=(i in table.highlight_rows)) for i, row in enumerate(table.rows)
    )
    totals_row = row_html(table.totals, bold=True) if table.totals else ""

    html_doc = f"""<!doctype html>
<html dir="{dir_attr}" lang="{"ar" if table.rtl else "en"}">
<head>
<meta charset="utf-8">
<style>
  @page {{ size: A4; margin: 16mm 12mm; }}
  body {{ font-family: "Noto Sans Arabic", "Noto Naskh Arabic", sans-serif; font-size: 10pt; color: #1a1a1a; }}
  h1 {{ font-size: 15pt; margin: 0 0 2pt 0; }}
  .company {{ font-size: 11pt; color: #444; margin: 0 0 1pt 0; }}
  .subtitle {{ font-size: 9pt; color: #666; margin: 0 0 10pt 0; }}
  table {{ width: 100%; border-collapse: collapse; }}
  th, td {{ border: 1px solid #ccc; padding: 4px 6px; font-variant-numeric: tabular-nums; }}
  th {{ background: #f2f2f2; font-weight: 700; }}
</style>
</head>
<body>
  <p class="company">{h(table.company_name)}</p>
  <h1>{h(table.title)}</h1>
  {f'<p class="subtitle">{h(table.subtitle)}</p>' if table.subtitle else ""}
  <table>
    <thead><tr>{header_cells}</tr></thead>
    <tbody>{body_rows}{totals_row}</tbody>
  </table>
</body>
</html>"""

    return HTML(string=html_doc).write_pdf()


def render_excel(table: ReportTable) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = table.title[:31] or "Report"
    ws.sheet_view.rightToLeft = table.rtl

    ws.append([table.company_name])
    ws.append([table.title])
    if table.subtitle:
        ws.append([table.subtitle])
    ws.append([])

    header_row_idx = ws.max_row + 1
    ws.append([c.label for c in table.columns])
    for cell in ws[header_row_idx]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right" if table.rtl else "left")

    for row in table.rows:
        ws.append(row)

    if table.totals:
        ws.append(table.totals)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

    for i, col in enumerate(table.columns, start=1):
        letter = get_column_letter(i)
        max_len = max(
            [len(col.label)] + [len(str(r[i - 1])) for r in table.rows if i - 1 < len(r)]
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
